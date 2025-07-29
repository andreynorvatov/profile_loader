import os
import glob
import warnings
import pandas as pd
from openpyxl import load_workbook

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

ARTIFACTS_DIR = 'artifacts'

def extract_summary_sheet(xlsx_path):
    """Извлекает таблицу с листа '8.Итоговый профиль+сравнение' начиная с C13."""
    print(f'Обработка: {xlsx_path}')
    wb = load_workbook(xlsx_path, data_only=True)
    if '8.Итоговый профиль+сравнение' not in wb.sheetnames:
        print(f'  Лист "8.Итоговый профиль+сравнение" не найден, пропускаем.')
        return None

    ws = wb['8.Итоговый профиль+сравнение']

    # Читаем диапазон C3:E24 сразу в DataFrame
    data = ws.iter_rows(min_row=3, max_row=24, min_col=3, max_col=5, values_only=True)
    df = pd.DataFrame(data)

    # Удаляем полностью пустые строки
    df.dropna(how='all', inplace=True)

    # Первую непустую строку используем как заголовок
    if not df.empty:
        df.columns = df.iloc[0]
        df = df.drop(df.index[0]).reset_index(drop=True)

    # Оставляем только нужные колонки: "Операция", "Интенсивность", "% распред. в профиле"
    df = df[['Операция', 'Интенсивность', '% распред. в профиле']]

    return df

def main():
    pattern = os.path.join(ARTIFACTS_DIR, 'profile_*.xlsx')
    files = glob.glob(pattern)
    if not files:
        print('Нет файлов profile_*.xlsx для обработки.')
        return

    all_dfs = []
    for file in files:
        df = extract_summary_sheet(file)
        if df is not None and not df.empty:
            # Добавляем колонку-источник (имя файла без расширения)
            df['_source'] = os.path.splitext(os.path.basename(file))[0]
            all_dfs.append(df)

    if not all_dfs:
        print('Нет данных для объединения.')
        return

    # Переименовываем колонки в каждой таблице, добавляя суффикс с именем файла
    renamed_dfs = []
    for df in all_dfs:
        source_name = df['_source'].iloc[0]
        df_renamed = df[['Операция', 'Интенсивность', '% распред. в профиле']].copy()
        df_renamed = df_renamed.rename(columns={
            'Интенсивность': f'{source_name}_Интенсивность',
            '% распред. в профиле': f'{source_name}_% распред. в профиле'
        })
        renamed_dfs.append(df_renamed)

    # Делаем join по колонке "Операция"
    merged = renamed_dfs[0]
    for df in renamed_dfs[1:]:
        merged = pd.merge(
            merged,
            df,
            on='Операция',
            how='outer'
        )

    output_path = os.path.join(ARTIFACTS_DIR, 'profile_summary_merged.csv')
    merged.to_csv(output_path, index=False, encoding='utf-8-sig')
    print(f'Объединённая таблица сохранена: {output_path}')

if __name__ == '__main__':
    main()
